import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# load xl workbook and return wb object
wb = xl.load_workbook('/Users/nathanielmention/pyxl/transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    # add a new column
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

# create instance of ref to hold values of 4th column
values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('/Users/nathanielmention/pyxl/transactions2.xlsx')
